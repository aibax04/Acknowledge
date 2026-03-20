import io
import logging
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_
from app.database import get_db
from app.models.user import User, UserRole
from app.models.attendance import Attendance, AttendanceStatus, AttendanceUpdateRequest
from app.models.holiday import Holiday
from app.models.leave import LeaveRequest, LeaveStatus
from app.routes.auth import get_current_user
from app.schemas.attendance_schema import (
    ClockInRequest, ClockOutRequest, AttendanceResponse,
    AttendanceUpdateRequestCreate, AttendanceUpdateRequestResponse, AttendanceUpdateReview,
    MarkAbsentRequest,
)
from datetime import datetime, date, time, timezone, timedelta
from typing import List, Optional

router = APIRouter(prefix="/attendance", tags=["attendance"])


def _normalize_office(office: Optional[str]) -> Optional[str]:
    """Normalize office value; accept legacy 'igen' as 'eigen'."""
    if not office:
        return office
    o = (office or "").strip().lower()
    if o == "igen":
        return "eigen"
    return office


def is_weekly_off(d: date, office: str) -> bool:
    """Check if a date is a weekly off for the given office."""
    office = _normalize_office(office) or office
    weekday = d.weekday()  # 0=Monday, 6=Sunday
    if office == "panscience":
        return weekday in (5, 6)  # Saturday, Sunday
    elif office == "eigen":
        return weekday == 6  # Sunday only
    return weekday == 6  # Default: Sunday off


async def get_holidays_for_date(db: AsyncSession, d: date, office: str) -> Optional[Holiday]:
    """Check if a date is a holiday for the given office."""
    office = _normalize_office(office) or office
    result = await db.execute(
        select(Holiday).filter(
            Holiday.date == d,
            (Holiday.office == office) | (Holiday.office == "both")
        )
    )
    return result.scalars().first()


@router.post("/clock-in")
async def clock_in(
    req: ClockInRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Clock in for today. Location is extracted from frontend."""
    now = datetime.now(timezone.utc)
    today = now.date()

    if not current_user.office:
        raise HTTPException(status_code=400, detail="Please set your office (Panscience/Eigen) in your profile first")

    # Check if it's a weekly off
    if is_weekly_off(today, current_user.office):
        raise HTTPException(status_code=400, detail="Today is a weekly off. No clock-in required.")

    # Check if it's a holiday
    holiday = await get_holidays_for_date(db, today, current_user.office)
    if holiday:
        raise HTTPException(status_code=400, detail=f"Today is a holiday: {holiday.title}. No clock-in required.")

    # Check working hours (9 AM to 7 PM IST = UTC+5:30)
    ist_offset = timedelta(hours=5, minutes=30)
    now_ist = now + ist_offset
    if now_ist.hour < 9 or now_ist.hour >= 19:
        raise HTTPException(status_code=400, detail="Clock-in is only allowed between 9:00 AM and 7:00 PM IST")

    # Check if already clocked in today
    result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == current_user.id,
            Attendance.date == today
        )
    )
    existing = result.scalars().first()
    if existing and existing.clock_in:
        raise HTTPException(status_code=400, detail="Already clocked in today")

    if existing:
        existing.clock_in = now
        existing.clock_in_lat = req.latitude
        existing.clock_in_lng = req.longitude
        existing.clock_in_address = req.address
        existing.status = AttendanceStatus.PRESENT
    else:
        attendance = Attendance(
            user_id=current_user.id,
            date=today,
            clock_in=now,
            clock_in_lat=req.latitude,
            clock_in_lng=req.longitude,
            clock_in_address=req.address,
            status=AttendanceStatus.PRESENT
        )
        db.add(attendance)

    await db.commit()
    return {"message": "Clocked in successfully", "time": now.isoformat()}


@router.post("/clock-out")
async def clock_out(
    req: ClockOutRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Clock out for today."""
    now = datetime.now(timezone.utc)
    today = now.date()

    # Check working hours
    ist_offset = timedelta(hours=5, minutes=30)
    now_ist = now + ist_offset
    if now_ist.hour < 9 or now_ist.hour >= 19:
        raise HTTPException(status_code=400, detail="Clock-out is only allowed between 9:00 AM and 7:00 PM IST")

    result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == current_user.id,
            Attendance.date == today
        )
    )
    existing = result.scalars().first()
    if not existing or not existing.clock_in:
        raise HTTPException(status_code=400, detail="You haven't clocked in today")
    if existing.clock_out:
        raise HTTPException(status_code=400, detail="Already clocked out today")

    existing.clock_out = now
    existing.clock_out_lat = req.latitude
    existing.clock_out_lng = req.longitude
    existing.clock_out_address = req.address
    await db.commit()

    return {"message": "Clocked out successfully", "time": now.isoformat()}


@router.get("/today")
async def get_today_attendance(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get today's attendance status for current user."""
    today = date.today()

    if not current_user.office:
        return {"status": "no_office", "message": "Please set your office first"}

    # Check weekly off
    if is_weekly_off(today, current_user.office):
        return {"status": "weekly_off", "message": "Weekly Off", "date": today.isoformat()}

    # Check holiday
    holiday = await get_holidays_for_date(db, today, current_user.office)
    if holiday:
        return {"status": "holiday", "message": f"Holiday: {holiday.title}", "date": today.isoformat()}

    # Check existing attendance record
    result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == current_user.id,
            Attendance.date == today
        )
    )
    record = result.scalars().first()

    if not record:
        return {"status": "not_clocked_in", "date": today.isoformat(), "clock_in": None, "clock_out": None}

    return {
        "status": record.status.value if record.status else "present",
        "date": today.isoformat(),
        "clock_in": record.clock_in.isoformat() if record.clock_in else None,
        "clock_out": record.clock_out.isoformat() if record.clock_out else None,
        "clock_in_address": record.clock_in_address,
        "clock_out_address": record.clock_out_address
    }


@router.get("/monthly")
async def get_monthly_attendance(
    year: int,
    month: int,
    user_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get monthly attendance for a user. If user_id not provided, gets current user's."""
    target_user_id = user_id or current_user.id

    # Only managers/seniors can view other users' attendance
    if target_user_id != current_user.id:
        if current_user.role not in (UserRole.MANAGER, UserRole.SENIOR):
            raise HTTPException(status_code=403, detail="Access denied")

    # Get the target user for office info
    if target_user_id == current_user.id:
        target_user = current_user
    else:
        result = await db.execute(select(User).filter(User.id == target_user_id))
        target_user = result.scalars().first()
        if not target_user:
            raise HTTPException(status_code=404, detail="User not found")

    office = _normalize_office(target_user.office) or "eigen"

    # Get date range
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)

    # Get all attendance records for the month
    result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == target_user_id,
            Attendance.date >= first_day,
            Attendance.date <= last_day
        )
    )
    records = {r.date: r for r in result.scalars().all()}

    # Get holidays for the month
    holiday_result = await db.execute(
        select(Holiday).filter(
            Holiday.date >= first_day,
            Holiday.date <= last_day,
            (Holiday.office == office) | (Holiday.office == "both")
        )
    )
    holidays = {h.date: h for h in holiday_result.scalars().all()}

    # Get approved leaves for the month
    leave_result = await db.execute(
        select(LeaveRequest).filter(
            LeaveRequest.user_id == target_user_id,
            LeaveRequest.status == LeaveStatus.APPROVED,
            LeaveRequest.start_date <= last_day,
            LeaveRequest.end_date >= first_day
        )
    )
    leaves = leave_result.scalars().all()
    leave_dates = set()
    for leave in leaves:
        d = max(leave.start_date, first_day)
        while d <= min(leave.end_date, last_day):
            leave_dates.add(d)
            d += timedelta(days=1)

    # Build attendance for each day. Order matters: holiday is checked before absent
    # so that no day that is a holiday for this office is ever marked absent.
    today = date.today()
    attendance_list = []
    d = first_day
    while d <= last_day:
        if d > today:
            attendance_list.append({
                "date": d.isoformat(),
                "status": "future",
                "clock_in": None,
                "clock_out": None
            })
        elif is_weekly_off(d, office):
            attendance_list.append({
                "date": d.isoformat(),
                "status": "weekly_off",
                "clock_in": None,
                "clock_out": None
            })
        elif d in holidays:
            attendance_list.append({
                "date": d.isoformat(),
                "status": "holiday",
                "holiday_name": holidays[d].title,
                "clock_in": None,
                "clock_out": None
            })
        elif d in leave_dates:
            attendance_list.append({
                "date": d.isoformat(),
                "status": "on_leave",
                "clock_in": None,
                "clock_out": None
            })
        elif d in records:
            rec = records[d]
            attendance_list.append({
                "date": d.isoformat(),
                "status": rec.status.value if rec.status else "present",
                "clock_in": rec.clock_in.isoformat() if rec.clock_in else None,
                "clock_out": rec.clock_out.isoformat() if rec.clock_out else None,
                "clock_in_address": rec.clock_in_address,
                "clock_out_address": rec.clock_out_address
            })
        else:
            attendance_list.append({
                "date": d.isoformat(),
                "status": "absent",
                "clock_in": None,
                "clock_out": None
            })
        d += timedelta(days=1)

    return {
        "user_id": target_user_id,
        "user_name": target_user.full_name,
        "office": office,
        "year": year,
        "month": month,
        "attendance": attendance_list
    }


@router.post("/mark-absent")
async def mark_absent(
    body: MarkAbsentRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Mark an employee as absent on a particular date. Manager/Senior only."""
    if current_user.role not in (UserRole.MANAGER, UserRole.SENIOR):
        raise HTTPException(status_code=403, detail="Only managers and directors can mark absent")

    user_id = body.user_id
    absent_date = body.absent_date
    today = date.today()
    if absent_date > today:
        raise HTTPException(status_code=400, detail="Cannot mark absent for a future date")

    result = await db.execute(select(User).filter(User.id == user_id))
    target_user = result.scalars().first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    office = _normalize_office(target_user.office) or "eigen"
    if is_weekly_off(absent_date, office):
        raise HTTPException(status_code=400, detail="That date is a weekly off for this employee's office")
    holiday = await get_holidays_for_date(db, absent_date, office)
    if holiday:
        raise HTTPException(status_code=400, detail=f"That date is a holiday: {holiday.title}")

    result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == user_id,
            Attendance.date == absent_date,
        )
    )
    existing = result.scalars().first()
    if existing:
        existing.clock_in = None
        existing.clock_out = None
        existing.status = AttendanceStatus.ABSENT
        existing.clock_in_address = None
        existing.clock_out_address = None
    else:
        att = Attendance(
            user_id=user_id,
            date=absent_date,
            clock_in=None,
            clock_out=None,
            status=AttendanceStatus.ABSENT,
        )
        db.add(att)
    await db.commit()
    return {"message": f"Marked {target_user.full_name} as absent on {absent_date.isoformat()}"}


@router.get("/export")
async def export_attendance_excel(
    year: int,
    month: int,
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export a user's monthly attendance as an Excel (.xlsx) file. Only managers/seniors."""
    if current_user.role not in (UserRole.MANAGER, UserRole.SENIOR):
        raise HTTPException(status_code=403, detail="Only managers and directors can export attendance")

    result = await db.execute(select(User).filter(User.id == user_id))
    target_user = result.scalars().first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    office = _normalize_office(target_user.office) or "eigen"
    first_day = date(year, month, 1)
    last_day = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year, 12, 31)

    att_result = await db.execute(
        select(Attendance).filter(
            Attendance.user_id == user_id,
            Attendance.date >= first_day,
            Attendance.date <= last_day
        )
    )
    records = {r.date: r for r in att_result.scalars().all()}

    hol_result = await db.execute(
        select(Holiday).filter(
            Holiday.date >= first_day,
            Holiday.date <= last_day,
            (Holiday.office == office) | (Holiday.office == "both")
        )
    )
    holidays = {h.date: h for h in hol_result.scalars().all()}

    leave_result = await db.execute(
        select(LeaveRequest).filter(
            LeaveRequest.user_id == user_id,
            LeaveRequest.status == LeaveStatus.APPROVED,
            LeaveRequest.start_date <= last_day,
            LeaveRequest.end_date >= first_day
        )
    )
    leaves = leave_result.scalars().all()
    leave_dates = set()
    for lv in leaves:
        d = max(lv.start_date, first_day)
        while d <= min(lv.end_date, last_day):
            leave_dates.add(d)
            d += timedelta(days=1)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logging.error("openpyxl not installed – falling back to CSV")
        raise HTTPException(status_code=500, detail="Excel library not available on server. Please ask admin to install openpyxl.")

    wb = Workbook()
    ws = wb.active
    month_name = date(year, month, 1).strftime("%B %Y")
    ws.title = month_name[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance Report – {target_user.full_name} – {month_name}"
    title_cell.font = Font(bold=True, size=14, color="111827")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    meta_cell = ws["A2"]
    meta_cell.value = f"Office: {office.capitalize()}  |  Exported by: {current_user.full_name}"
    meta_cell.font = Font(size=10, color="6B7280")
    meta_cell.alignment = Alignment(horizontal="center")

    headers = ["Date", "Day", "Status", "Clock In (IST)", "Clock Out (IST)", "Hours"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ist = timedelta(hours=5, minutes=30)
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    status_fills = {
        "present": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "absent": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "weekly_off": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        "holiday": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        "on_leave": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "future": PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"),
    }
    today = date.today()
    row = 5
    present_count = absent_count = leave_count = holiday_count = woff_count = 0
    total_hours = 0.0

    d = first_day
    while d <= last_day:
        day_name = day_names[d.weekday()]
        if d > today:
            status = "Future"
            ci_str = co_str = hrs_str = ""
        elif is_weekly_off(d, office):
            status = "Weekly Off"
            ci_str = co_str = hrs_str = ""
            woff_count += 1
        elif d in holidays:
            status = f"Holiday – {holidays[d].title}"
            ci_str = co_str = hrs_str = ""
            holiday_count += 1
        elif d in leave_dates:
            status = "On Leave"
            ci_str = co_str = hrs_str = ""
            leave_count += 1
        elif d in records:
            rec = records[d]
            if rec.status == AttendanceStatus.ABSENT:
                status = "Absent"
                ci_str = co_str = hrs_str = ""
                absent_count += 1
            else:
                status = "Present"
                ci_str = (rec.clock_in + ist).strftime("%I:%M %p") if rec.clock_in else "-"
                co_str = (rec.clock_out + ist).strftime("%I:%M %p") if rec.clock_out else "-"
                hrs = ""
                if rec.clock_in and rec.clock_out:
                    delta = (rec.clock_out - rec.clock_in).total_seconds() / 3600
                    hrs = f"{delta:.1f}"
                    total_hours += delta
                hrs_str = hrs
                present_count += 1
        else:
            status = "Absent"
            ci_str = co_str = hrs_str = ""
            absent_count += 1

        ws.cell(row=row, column=1, value=d.strftime("%d %b %Y")).border = thin_border
        ws.cell(row=row, column=2, value=day_name).border = thin_border
        status_cell = ws.cell(row=row, column=3, value=status)
        status_cell.border = thin_border
        ws.cell(row=row, column=4, value=ci_str).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=co_str).border = thin_border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=hrs_str).border = thin_border
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")

        status_key = status.lower().split(" –")[0].replace(" ", "_")
        fill = status_fills.get(status_key)
        if fill:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = fill

        d += timedelta(days=1)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    summary_cell = ws.cell(row=row, column=1)
    summary_cell.value = f"Summary: Present {present_count} | Absent {absent_count} | Leave {leave_count} | Holiday {holiday_count} | Weekly Off {woff_count} | Total Hours {total_hours:.1f}"
    summary_cell.font = Font(bold=True, size=10, color="374151")
    summary_cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = target_user.full_name.replace(" ", "_")
    filename = f"Attendance_{safe_name}_{month_name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================
# ATTENDANCE UPDATE REQUESTS
# ============================================

@router.post("/update-request")
async def create_update_request(
    req: AttendanceUpdateRequestCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Request to update attendance for a past date."""
    today = date.today()
    if req.date >= today:
        raise HTTPException(status_code=400, detail="Can only request updates for past dates")

    # Parse the clock-in/out times
    requested_clock_in = None
    requested_clock_out = None
    if req.requested_clock_in:
        try:
            t = datetime.fromisoformat(req.requested_clock_in)
            requested_clock_in = t.replace(tzinfo=timezone.utc) if t.tzinfo is None else t
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid clock-in time format")
    if req.requested_clock_out:
        try:
            t = datetime.fromisoformat(req.requested_clock_out)
            requested_clock_out = t.replace(tzinfo=timezone.utc) if t.tzinfo is None else t
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid clock-out time format")

    # Verify manager exists and is actually a manager
    mgr_result = await db.execute(select(User).filter(User.id == req.manager_id))
    manager = mgr_result.scalars().first()
    if not manager or manager.role not in (UserRole.MANAGER, UserRole.SENIOR):
        raise HTTPException(status_code=400, detail="Invalid manager selected")

    update_request = AttendanceUpdateRequest(
        user_id=current_user.id,
        date=req.date,
        requested_clock_in=requested_clock_in,
        requested_clock_out=requested_clock_out,
        reason=req.reason,
        manager_id=req.manager_id,
        status="pending"
    )
    db.add(update_request)
    await db.commit()
    await db.refresh(update_request)

    return {"message": "Update request submitted successfully", "id": update_request.id}


@router.get("/update-requests/pending", response_model=List[AttendanceUpdateRequestResponse])
async def get_pending_update_requests(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get pending attendance update requests for current manager."""
    if current_user.role not in (UserRole.MANAGER, UserRole.SENIOR):
        raise HTTPException(status_code=403, detail="Only managers/seniors can view update requests")

    result = await db.execute(
        select(AttendanceUpdateRequest).filter(
            AttendanceUpdateRequest.manager_id == current_user.id,
            AttendanceUpdateRequest.status == "pending"
        ).order_by(AttendanceUpdateRequest.created_at.desc())
    )
    requests = result.scalars().all()

    response = []
    for r in requests:
        user_result = await db.execute(select(User).filter(User.id == r.user_id))
        user = user_result.scalars().first()
        response.append(AttendanceUpdateRequestResponse(
            id=r.id,
            user_id=r.user_id,
            date=r.date,
            requested_clock_in=r.requested_clock_in,
            requested_clock_out=r.requested_clock_out,
            reason=r.reason,
            manager_id=r.manager_id,
            status=r.status,
            reviewer_notes=r.reviewer_notes,
            created_at=r.created_at,
            reviewed_at=r.reviewed_at,
            user_name=user.full_name if user else None,
            manager_name=current_user.full_name
        ))

    return response


@router.get("/my-update-requests")
async def get_my_update_requests(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get current user's attendance update requests."""
    result = await db.execute(
        select(AttendanceUpdateRequest).filter(
            AttendanceUpdateRequest.user_id == current_user.id
        ).order_by(AttendanceUpdateRequest.created_at.desc())
    )
    requests = result.scalars().all()

    response = []
    for r in requests:
        mgr_result = await db.execute(select(User).filter(User.id == r.manager_id))
        mgr = mgr_result.scalars().first()
        response.append({
            "id": r.id,
            "date": r.date.isoformat(),
            "requested_clock_in": r.requested_clock_in.isoformat() if r.requested_clock_in else None,
            "requested_clock_out": r.requested_clock_out.isoformat() if r.requested_clock_out else None,
            "reason": r.reason,
            "manager_name": mgr.full_name if mgr else "Unknown",
            "status": r.status,
            "reviewer_notes": r.reviewer_notes,
            "created_at": r.created_at.isoformat()
        })

    return response


@router.put("/update-requests/{request_id}/review")
async def review_update_request(
    request_id: int,
    review: AttendanceUpdateReview,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Approve or reject an attendance update request."""
    if current_user.role not in (UserRole.MANAGER, UserRole.SENIOR):
        raise HTTPException(status_code=403, detail="Only managers/seniors can review requests")

    result = await db.execute(
        select(AttendanceUpdateRequest).filter(AttendanceUpdateRequest.id == request_id)
    )
    req = result.scalars().first()
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req.manager_id != current_user.id:
        raise HTTPException(status_code=403, detail="This request is not assigned to you")
    if req.status != "pending":
        raise HTTPException(status_code=400, detail="This request has already been reviewed")

    if review.status not in ("approved", "rejected"):
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")

    req.status = review.status
    req.reviewer_notes = review.reviewer_notes
    req.reviewed_at = datetime.now(timezone.utc)

    if review.status == "approved":
        # Update or create attendance record
        att_result = await db.execute(
            select(Attendance).filter(
                Attendance.user_id == req.user_id,
                Attendance.date == req.date
            )
        )
        attendance = att_result.scalars().first()
        if attendance:
            if req.requested_clock_in:
                attendance.clock_in = req.requested_clock_in
            if req.requested_clock_out:
                attendance.clock_out = req.requested_clock_out
            attendance.status = AttendanceStatus.PRESENT
        else:
            attendance = Attendance(
                user_id=req.user_id,
                date=req.date,
                clock_in=req.requested_clock_in,
                clock_out=req.requested_clock_out,
                status=AttendanceStatus.PRESENT
            )
            db.add(attendance)

    await db.commit()
    return {"message": f"Request {review.status} successfully"}


@router.get("/managers")
async def get_managers_list(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get list of managers for attendance update request dropdown."""
    result = await db.execute(
        select(User).filter(
            User.role.in_([UserRole.MANAGER, UserRole.SENIOR]),
            User.is_active == True
        )
    )
    managers = result.scalars().all()
    return [{"id": m.id, "full_name": m.full_name, "role": m.role.value} for m in managers]
